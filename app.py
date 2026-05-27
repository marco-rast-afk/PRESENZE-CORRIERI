import streamlit as st
import pandas as pd
from fpdf import FPDF
import io
import json
import requests
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Gestione Presenze Corrieri", layout="wide", initial_sidebar_state="expanded")

# ─────────────────────────────────────────────────────────────────────────────
# FIX #1 – SCROLL PRESERVATION
# Inietta JS che memorizza la posizione Y prima di ogni re-render e la ripristina
# dopo. Streamlit ricarica l'intera pagina ad ogni interazione; questo snippet
# aggancia window.onbeforeunload per salvare scrollY in sessionStorage e
# un MutationObserver sul body per ripristinarlo non appena il DOM è pronto.
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <script>
    (function() {
        const KEY = 'st_scroll_pos';

        // Salva posizione prima del reload
        window.addEventListener('beforeunload', () => {
            sessionStorage.setItem(KEY, window.scrollY);
        });

        // Ripristina posizione dopo che Streamlit ha ri-renderizzato il DOM
        function restoreScroll() {
            const saved = sessionStorage.getItem(KEY);
            if (saved !== null) {
                window.scrollTo(0, parseInt(saved, 10));
                sessionStorage.removeItem(KEY);
            }
        }

        // Osserva il body: appena Streamlit termina il rendering, ripristina
        const observer = new MutationObserver(() => {
            restoreScroll();
        });
        observer.observe(document.body, { childList: true, subtree: true });

        // Fallback: ripristina anche al normale DOMContentLoaded
        document.addEventListener('DOMContentLoaded', restoreScroll);
    })();
    </script>
    """,
    unsafe_allow_html=True,
)

# CSS: celle compatte, griglia massimizzata, padding ridotto
st.markdown("""
<style>
/* Riduci padding pagina per massimizzare spazio griglia */
.block-container {
    padding-top: 0.5rem !important;
    padding-bottom: 0.5rem !important;
    padding-left: 1rem !important;
    padding-right: 1rem !important;
    max-width: 100% !important;
}

/* Celle griglia piu basse */
div[data-testid="stDataFrame"] td,
div[data-testid="stDataFrame"] th {
    padding-top: 2px !important;
    padding-bottom: 2px !important;
    line-height: 1.2 !important;
    font-size: 0.82rem !important;
    white-space: nowrap !important;
}

/* Altezza riga minima */
div[data-testid="stDataFrame"] tr {
    min-height: 24px !important;
    height: 24px !important;
}

/* Data editor celle compatte */
div[data-testid="stDataEditor"] td,
div[data-testid="stDataEditor"] th {
    padding-top: 2px !important;
    padding-bottom: 2px !important;
    line-height: 1.2 !important;
    font-size: 0.82rem !important;
}

div[data-testid="stDataEditor"] tr {
    min-height: 24px !important;
    height: 24px !important;
}

/* Riduci spazio date input e altri widget sopra la griglia */
div[data-testid="stDateInput"] {
    margin-bottom: 4px !important;
}

/* Riduci altezza header senza nasconderlo (nasconderlo blocca il toggle sidebar) */
header[data-testid="stHeader"] {
    height: 2.5rem !important;
    min-height: 2.5rem !important;
}

/* Riduci margini tra elementi */
div[data-testid="stVerticalBlock"] > div {
    gap: 0.25rem !important;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SUPABASE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
SUPABASE_URL = st.secrets["SUPABASE_URL"].rstrip("/")
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
SB_HEADERS = {
    "apikey":        SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type":  "application/json",
    "Prefer":        "return=minimal",
}

def _sb_url(tabella: str) -> str:
    return f"{SUPABASE_URL}/rest/v1/{tabella}"

# Dizionario per convertire il numero del mese nel nome in italiano maiuscolo
MESI_ITA = {
    1: "GENNAIO", 2: "FEBBRAIO", 3: "MARZO", 4: "APRILE", 5: "MAGGIO", 6: "GIUGNO",
    7: "LUGLIO", 8: "AGOSTO", 9: "SETTEMBRE", 10: "OTTOBRE", 11: "NOVEMBRE", 12: "DICEMBRE"
}

def parse_data_ita(s):
    """Converte stringa 'G MESE AAAA' in datetime.date. Ritorna None se fallisce."""
    try:
        mesi_inv = {v: k for k, v in MESI_ITA.items()}
        parti = str(s).strip().split()
        if len(parti) == 3:
            g, m, a = int(parti[0]), mesi_inv.get(parti[1].upper(), 0), int(parti[2])
            if m:
                return datetime(a, m, g).date()
    except Exception:
        pass
    return None

# ─────────────────────────────────────────────────────────────────────────────
# SALVATAGGIO / CARICAMENTO SUPABASE
# ─────────────────────────────────────────────────────────────────────────────
def _sb_truncate(tabella: str):
    """Svuota la tabella via SQL RPC (unico modo affidabile in Supabase)."""
    requests.post(
        f"{SUPABASE_URL}/rest/v1/rpc/truncate_table",
        headers={**SB_HEADERS, "Content-Type": "application/json"},
        data=json.dumps({"table_name": tabella})
    )

def _sb_inserisci(tabella: str, records: list):
    """Inserisce una lista di record nella tabella."""
    if not records:
        return
    puliti = [{k: v for k, v in r.items() if k != "id"} for r in records]
    requests.post(
        _sb_url(tabella),
        headers={**SB_HEADERS, "Content-Type": "application/json"},
        data=json.dumps(puliti)
    )

def _sb_upsert(tabella: str, records: list):
    """Svuota la tabella e reinserisce i record aggiornati."""
    _sb_truncate(tabella)
    _sb_inserisci(tabella, records)

def _sb_leggi(tabella: str) -> list:
    """Legge TUTTI i record di una tabella Supabase usando paginazione a blocchi.
    Supabase REST API restituisce al massimo 1000 righe per richiesta per default.
    """
    PAGE_SIZE = 1000
    tutti = []
    offset = 0
    headers_pag = {
        **SB_HEADERS,
        "Range-Unit": "items",
        "Prefer":     "count=none",
    }
    while True:
        fine = offset + PAGE_SIZE - 1
        headers_pag["Range"] = f"{offset}-{fine}"
        r = requests.get(
            _sb_url(tabella) + "?order=id.asc",
            headers=headers_pag
        )
        if r.status_code not in (200, 206):
            break
        blocco = r.json()
        if not blocco:
            break
        tutti.extend(blocco)
        if len(blocco) < PAGE_SIZE:
            break
        offset += PAGE_SIZE
    return [{k: v for k, v in row.items() if k != "id"} for row in tutti]

def salva_database_json():
    """Salva tutti i DataFrame su Supabase."""
    try:
        _sb_upsert("furgoni",            st.session_state.furgoni.to_dict(orient="records"))
        _sb_upsert("anagrafica_corrieri", st.session_state.anagrafica_corrieri.to_dict(orient="records"))
        _sb_upsert("responsabili",        st.session_state.responsabili.to_dict(orient="records"))
        _sb_upsert("stato_giornaliero",   st.session_state.stato_giornaliero.to_dict(orient="records"))
        _sb_upsert("storico_presenze",    st.session_state.storico_presenze.to_dict(orient="records"))
        st.sidebar.success("Salvato su Supabase!")
        return True
    except Exception as e:
        st.error(f"Errore salvataggio Supabase: {e}")
        return False

def carica_database_json():
    """Carica i dati da Supabase."""
    try:
        furgoni   = _sb_leggi("furgoni")
        corrieri  = _sb_leggi("anagrafica_corrieri")
        resp      = _sb_leggi("responsabili")
        giorno    = _sb_leggi("stato_giornaliero")
        storico   = _sb_leggi("storico_presenze")
        if furgoni or corrieri:
            st.session_state.furgoni             = pd.DataFrame(furgoni)
            df_anagrafica                        = pd.DataFrame(corrieri) if corrieri else pd.DataFrame()
            st.session_state.anagrafica_corrieri = df_anagrafica
            st.session_state.responsabili        = pd.DataFrame(resp)    if resp     else pd.DataFrame()
            st.session_state.storico_presenze    = pd.DataFrame(storico) if storico  else pd.DataFrame()

            # Ricostruisce stato_giornaliero fondendo con anagrafica se mancano COGNOME/NOME/CELLULARE/GIRO_FISSO
            df_giorno = pd.DataFrame(giorno) if giorno else pd.DataFrame()
            colonne_anag = ["COGNOME", "NOME", "CELLULARE", "GIRO_FISSO"]
            if not df_giorno.empty and not df_anagrafica.empty:
                mancanti = [c for c in colonne_anag if c not in df_giorno.columns]
                if mancanti:
                    df_giorno = df_anagrafica.merge(
                        df_giorno.drop(columns=[c for c in colonne_anag if c in df_giorno.columns], errors="ignore"),
                        left_index=True, right_index=True, how="left"
                    )
            elif df_giorno.empty and not df_anagrafica.empty:
                df_giorno = df_anagrafica.copy()
                df_giorno["STATO"]         = "Presente (Giro Fisso)"
                df_giorno["GIRO_SUPPORTO"] = ""
                df_giorno["MEZZO"]         = "Nessuno"
                df_giorno["KM_INIZIO"]     = 0
                df_giorno["NOTE"]          = ""
            st.session_state.stato_giornaliero = df_giorno
            return True
        return False
    except Exception as e:
        st.error(f"Errore caricamento Supabase: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# INIZIALIZZAZIONE DATABASE
# ─────────────────────────────────────────────────────────────────────────────
if 'database_caricato' not in st.session_state:
    if carica_database_json():
        st.session_state.database_caricato = True
    else:
        st.session_state.furgoni = pd.DataFrame([
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "EP800MZ", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "EX184SW", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "EX804EN", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "EX805EN", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "FA395MK", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "FD357DK", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "FJ898TP", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "FN446DZ", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "FP750ED", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "FR953KJ", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "FX735HG", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "FX845YP", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "FX883NS", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GC957VZ", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GF235BK", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GF237BK", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GF238BK", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GF239BK", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GF298BK", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GG111WM", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GG112WM", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GG149WM", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GG150WM", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GG151WM", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GG184WM", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GG187WM", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GG188WM", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GH880BE", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GL849VF", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GL850VF", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GL851VF", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GN724ES", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GN728ES", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GP529LX", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GR250RF", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GR256RF", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GR450EF", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GR452EF", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GR474EF", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GR964MV", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GS386TT", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GS387TT", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GS554WM", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GS556WM", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GS557WM", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GT874LE", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GX245HJ", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GX250HJ", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GX300XE", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GX301XE", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GX322TJ", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GX358HJ", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GX363HJ", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GX582FK", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GX831FN", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GX837FN", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GX852FN", "DISPONIBILE": "SI"},
            {"MARCA": "Fiat",     "MODELLO": "Ducato",   "TIPO": "Furgone", "TARGA": "GY571FV", "DISPONIBILE": "SI"},
            # FIX #5 – MARCA corretta per i Mercedes
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HA062CS", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HA385CY", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HA386CY", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HA902CR", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HB659CE", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HB662CE", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HB683CE", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HB783TS", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HB784TS", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HB785TS", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HB786TS", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HB787TS", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HB788TS", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HB789TS", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HB790TS", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HB792TS", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HB793TS", "DISPONIBILE": "SI"},
            {"MARCA": "Mercedes", "MODELLO": "Sprinter", "TIPO": "Furgone", "TARGA": "HC627CS", "DISPONIBILE": "SI"},
        ])

        st.session_state.anagrafica_corrieri = pd.DataFrame([
            {"COGNOME": "CROCI",         "NOME": "MARINO",        "CELLULARE": "3314509080", "GIRO_FISSO": "1"},
            {"COGNOME": "VINCIGUERRA",   "NOME": "ANGELO",        "CELLULARE": "3351696753", "GIRO_FISSO": "2"},
            {"COGNOME": "D ANGELO",      "NOME": "SALVATORE",     "CELLULARE": "3881132883", "GIRO_FISSO": "3"},
            {"COGNOME": "MARCIANO",      "NOME": "ANTONIO",       "CELLULARE": "3489292359", "GIRO_FISSO": "4"},
            {"COGNOME": "CAPUTO",        "NOME": "OVIDIO",        "CELLULARE": "3385277033", "GIRO_FISSO": "5"},
            {"COGNOME": "AINIS",         "NOME": "CIRO",          "CELLULARE": "3891618386", "GIRO_FISSO": "6"},
            {"COGNOME": "CAPPELLA",      "NOME": "GIUSEPPE",      "CELLULARE": "3427598835", "GIRO_FISSO": "7"},
            {"COGNOME": "BEN CHATTI",    "NOME": "MARTIN",        "CELLULARE": "3451303466", "GIRO_FISSO": "8"},
            {"COGNOME": "POMILIA",       "NOME": "PIETRO",        "CELLULARE": "3382107031", "GIRO_FISSO": "9"},
            {"COGNOME": "CAPUTO",        "NOME": "GIANLUCA",      "CELLULARE": "3391178350", "GIRO_FISSO": "10"},
            {"COGNOME": "CAIONI",        "NOME": "MANUEL",        "CELLULARE": "3703349078", "GIRO_FISSO": "11"},
            {"COGNOME": "MANZO",         "NOME": "GIAMPAOLO",     "CELLULARE": "3384187899", "GIRO_FISSO": "12"},
            {"COGNOME": "XHANI",         "NOME": "ORGES",         "CELLULARE": "3515960708", "GIRO_FISSO": "13"},
            {"COGNOME": "CISSE",         "NOME": "THIERNO",       "CELLULARE": "3513434477", "GIRO_FISSO": "14"},
            {"COGNOME": "CAPPONI",       "NOME": "MASSIMILIANO",  "CELLULARE": "3206274943", "GIRO_FISSO": "15"},
            {"COGNOME": "MERCURI",       "NOME": "SIMONE",        "CELLULARE": "3483624304", "GIRO_FISSO": "16"},
            {"COGNOME": "FELICIANI",     "NOME": "SAMUELE",       "CELLULARE": "3459494393", "GIRO_FISSO": "17"},
            {"COGNOME": "CALUGI",        "NOME": "ALESSANDRO",    "CELLULARE": "3467894592", "GIRO_FISSO": "18"},
            {"COGNOME": "D'ANTONIO",     "NOME": "GIANFRANCO",    "CELLULARE": "3791766322", "GIRO_FISSO": "19"},
            {"COGNOME": "SPACCASASSI",   "NOME": "MARIO",         "CELLULARE": "3298424239", "GIRO_FISSO": "20"},
            {"COGNOME": "MACRILLANTE",   "NOME": "LORENZO",       "CELLULARE": "3396085295", "GIRO_FISSO": "21"},
            {"COGNOME": "CAPPELLI",      "NOME": "SILVIO",        "CELLULARE": "3314578329", "GIRO_FISSO": "22"},
            {"COGNOME": "SQUILLACE",     "NOME": "MICHELE",       "CELLULARE": "3404966459", "GIRO_FISSO": "23"},
            {"COGNOME": "CAPRIOTTI",     "NOME": "CRISTIANO",     "CELLULARE": "3452137376", "GIRO_FISSO": "24"},
            {"COGNOME": "VINCIGUERRA",   "NOME": "VINCENZO",      "CELLULARE": "3348361930", "GIRO_FISSO": "25"},
            {"COGNOME": "AMANDONICO",    "NOME": "DARIO",         "CELLULARE": "3477784701", "GIRO_FISSO": "26"},
            {"COGNOME": "DI PIETRO",     "NOME": "SIMONE",        "CELLULARE": "3484587690", "GIRO_FISSO": "27"},
            {"COGNOME": "LUCIANI",       "NOME": "GIOVANNI",      "CELLULARE": "3282063754", "GIRO_FISSO": "28"},
            {"COGNOME": "DE SIMONE",     "NOME": "ANTONIO",       "CELLULARE": "3472316522", "GIRO_FISSO": "29"},
            {"COGNOME": "MARCONI",       "NOME": "ANDREA",        "CELLULARE": "3381092067", "GIRO_FISSO": "30"},
            {"COGNOME": "PETROCCHI",     "NOME": "ALESSANDRO",    "CELLULARE": "3204269922", "GIRO_FISSO": "31"},
            {"COGNOME": "MANAZZA",       "NOME": "GIUSEPPE",      "CELLULARE": "3926854314", "GIRO_FISSO": "32"},
            {"COGNOME": "SCARPA",        "NOME": "CRESCENZO",     "CELLULARE": "3494993238", "GIRO_FISSO": "33"},
            {"COGNOME": "FEBO",          "NOME": "DANIELE",       "CELLULARE": "3297074905", "GIRO_FISSO": "34"},
            {"COGNOME": "DIENG",         "NOME": "MAYPE",         "CELLULARE": "3203619657", "GIRO_FISSO": "35"},
            {"COGNOME": "MARCONI",       "NOME": "ALESSIO",       "CELLULARE": "3931574397", "GIRO_FISSO": "36"},
            {"COGNOME": "CORCIONE",      "NOME": "GIUSEPPE",      "CELLULARE": "3203437844", "GIRO_FISSO": "37"},
            {"COGNOME": "PAVAN",         "NOME": "ALESSANDRO",    "CELLULARE": "3515953354", "GIRO_FISSO": "38"},
            {"COGNOME": "CAMUSO",        "NOME": "LUCA",          "CELLULARE": "3511244208", "GIRO_FISSO": "39"},
            {"COGNOME": "NICCOLINI",     "NOME": "ANDREA",        "CELLULARE": "3202732824", "GIRO_FISSO": "40"},
            {"COGNOME": "VINOTTI",       "NOME": "SONNY",         "CELLULARE": "3313800865", "GIRO_FISSO": "41"},
            {"COGNOME": "SCIARRA",       "NOME": "ALESSANDRO",    "CELLULARE": "3515006439", "GIRO_FISSO": "42"},
            {"COGNOME": "ENOW",          "NOME": "SOLOMON",       "CELLULARE": "3930180387", "GIRO_FISSO": "43"},
            {"COGNOME": "GIACOMELLI",    "NOME": "PAOLO",         "CELLULARE": "3298884097", "GIRO_FISSO": "201"},
            {"COGNOME": "SPALAZZI",      "NOME": "BENEDETTA",     "CELLULARE": "3341971927", "GIRO_FISSO": "202"},
            {"COGNOME": "FERRAMINI",     "NOME": "PANCRAZIO",     "CELLULARE": "3408660269", "GIRO_FISSO": "203"},
            {"COGNOME": "OREFICE",       "NOME": "SALVATORE",     "CELLULARE": "3899012716", "GIRO_FISSO": "204"},  # FIX typo "SALVAVORE"
            {"COGNOME": "SALVI",         "NOME": "GIANLUCA",      "CELLULARE": "3289122867", "GIRO_FISSO": "205"},
            {"COGNOME": "GIULIANI",      "NOME": "LUIGI",         "CELLULARE": "3318356858", "GIRO_FISSO": "206"},
            {"COGNOME": "TOMMARELLI",    "NOME": "ALFONSO",       "CELLULARE": "3478548756", "GIRO_FISSO": "207"},
            {"COGNOME": "NISTOR",        "NOME": "BOGDAN",        "CELLULARE": "3293436637", "GIRO_FISSO": "208"},
            {"COGNOME": "DE MINICIS",    "NOME": "GIULIANO",      "CELLULARE": "3932174950", "GIRO_FISSO": "209"},
            {"COGNOME": "PERILLO",       "NOME": "VINCENZO",      "CELLULARE": "3349059267", "GIRO_FISSO": "210"},
            {"COGNOME": "BORGIA",        "NOME": "FRANCESCO",     "CELLULARE": "3280347303", "GIRO_FISSO": "211"},
            {"COGNOME": "VESPERINI",     "NOME": "ROBERTO",       "CELLULARE": "3203215836", "GIRO_FISSO": "212"},
            {"COGNOME": "PROTASI",       "NOME": "FILIPPO",       "CELLULARE": "3490626638", "GIRO_FISSO": "213"},
            {"COGNOME": "DE MATTEIS",    "NOME": "ALESSIA",       "CELLULARE": "3490566977", "GIRO_FISSO": "214"},
            {"COGNOME": "PELLE",         "NOME": "EUGENIO",       "CELLULARE": "3272314445", "GIRO_FISSO": "215"},
            {"COGNOME": "COSTANTINI",    "NOME": "FABRIZIO",      "CELLULARE": "3355869387", "GIRO_FISSO": "216"},
            {"COGNOME": "MASCITTI",      "NOME": "CLAUDIO",       "CELLULARE": "3404501899", "GIRO_FISSO": "217"},
            {"COGNOME": "TRANQUILLI",    "NOME": "SIMONE",        "CELLULARE": "3807991216", "GIRO_FISSO": "218"},
            {"COGNOME": "MARUSCO",       "NOME": "DAVIDE",        "CELLULARE": "3296488555", "GIRO_FISSO": "219"},
            {"COGNOME": "MASSICCI",      "NOME": "EMILIE",        "CELLULARE": "3896382624", "GIRO_FISSO": "220"},
            {"COGNOME": "BOFFINI",       "NOME": "VALERIO",       "CELLULARE": "3703393921", "GIRO_FISSO": "221"},
            {"COGNOME": "VINCIGUERRA",   "NOME": "ROSARIO",       "CELLULARE": "3888818725", "GIRO_FISSO": "222"},
            {"COGNOME": "TRANQUILLI",    "NOME": "JESSICA",       "CELLULARE": "3466767960", "GIRO_FISSO": "223"},
            {"COGNOME": "GERMINI",       "NOME": "MASSIMO",       "CELLULARE": "3381174017", "GIRO_FISSO": "224"},
            {"COGNOME": "FELICIANI",     "NOME": "MATTEO",        "CELLULARE": "3459494394", "GIRO_FISSO": "225"},
            {"COGNOME": "PERSICO",       "NOME": "GIADA",         "CELLULARE": "3500750399", "GIRO_FISSO": "226"},
            {"COGNOME": "MATJA",         "NOME": "KLAUDIO",       "CELLULARE": "3291692911", "GIRO_FISSO": "227"},
            {"COGNOME": "ROTUNNO",       "NOME": "DARIO",         "CELLULARE": "3387710023", "GIRO_FISSO": "228"},
            {"COGNOME": "LANZA",         "NOME": "CHRISTIAN",     "CELLULARE": "3881680533", "GIRO_FISSO": "230"},
            {"COGNOME": "RUZZINI",       "NOME": "ROBERTO",       "CELLULARE": "3283797115", "GIRO_FISSO": "501"},
            {"COGNOME": "STANGONI",      "NOME": "MARCO",         "CELLULARE": "3319145878", "GIRO_FISSO": "503"},
            {"COGNOME": "ECHEZURIA",     "NOME": "CARLOS",        "CELLULARE": "3317412770", "GIRO_FISSO": "508"},
            {"COGNOME": "ROSCIOLI",      "NOME": "PIERFRANCESCO", "CELLULARE": "3454280718", "GIRO_FISSO": "510"},
            {"COGNOME": "SPAGNOLI",      "NOME": "ANDREA",        "CELLULARE": "3756760553", "GIRO_FISSO": "511"},
            {"COGNOME": "MEDICO",        "NOME": "PIERPAOLO",     "CELLULARE": "3298537578", "GIRO_FISSO": "513"},
            {"COGNOME": "MUHAMMAD",      "NOME": "YOUNAS",        "CELLULARE": "3463618752", "GIRO_FISSO": "514"},
            {"COGNOME": "IAVARONE",      "NOME": "GIUSTINO",      "CELLULARE": "3488701779", "GIRO_FISSO": "515"},
            {"COGNOME": "CATALANO",      "NOME": "SILVIO",        "CELLULARE": "3534763452", "GIRO_FISSO": "517"},
            {"COGNOME": "DI GIROLAMO",   "NOME": "ALESSANDRO",    "CELLULARE": "3802543375", "GIRO_FISSO": "518"},
        ])

        st.session_state.responsabili = pd.DataFrame([
            {"COGNOME": "ROSSI", "NOME": "LUIGI",  "RUOLO": "Responsabile Logistica"},
            {"COGNOME": "VERDI", "NOME": "MARCO",  "RUOLO": "Supervisore di Turno"}
        ])

        df_giorno = st.session_state.anagrafica_corrieri.copy()
        df_giorno["STATO"] = "Presente (Giro Fisso)"
        df_giorno["GIRO_SUPPORTO"] = ""
        df_giorno["MEZZO"] = "Nessuno"
        df_giorno["KM_INIZIO"] = 0
        df_giorno["NOTE"] = ""
        st.session_state.stato_giornaliero = df_giorno
        st.session_state.storico_presenze = pd.DataFrame(columns=[
            "DATA", "COGNOME", "NOME", "GIRO_FISSO", "STATO",
            "MEZZO", "KM_INIZIO", "KM_FINE", "KM_PERCORSI", "NOTE"
        ])
        st.session_state.database_caricato = True

# ─────────────────────────────────────────────────────────────────────────────
# MENU DI NAVIGAZIONE E BOTTONE SALVA FISSO NELLA SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
# Inizializza chiavi export in session_state per evitare AttributeError
for _k in ["excel_sidebar_data", "excel_sidebar_nome", "pdf_sidebar_data", "pdf_sidebar_nome"]:
    if _k not in st.session_state:
        st.session_state[_k] = None

# Inizializza storico_presenze se non ancora presente (es. DB precedente senza questa tabella)
if "storico_presenze" not in st.session_state:
    st.session_state.storico_presenze = pd.DataFrame(columns=[
        "DATA", "COGNOME", "NOME", "GIRO_FISSO", "STATO",
        "MEZZO", "KM_INIZIO", "KM_FINE", "KM_PERCORSI", "NOTE"
    ])

# Assicura che stato_giornaliero abbia TUTTE le colonne attese
# (colonne mancanti si verificano quando il DB è stato creato con una versione precedente)
_colonne_default = {
    "STATO":         "Presente (Giro Fisso)",
    "GIRO_SUPPORTO": "",
    "MEZZO":         "Nessuno",
    "KM_INIZIO":     0,
    "NOTE":          "",
}
for _col, _default in _colonne_default.items():
    if _col not in st.session_state.stato_giornaliero.columns:
        st.session_state.stato_giornaliero[_col] = _default

menu = ["📋 Tabellone Presenze", "🚐 Anagrafica Furgoni", "👥 Anagrafica Personale", "📊 Storico & Furgoni"]
scelta = st.sidebar.selectbox("Navigazione", menu)

st.sidebar.markdown("---")
st.sidebar.subheader("💾 Salvataggio Dati")
if st.sidebar.button("💾 SALVA SU DATABASE", use_container_width=True, type="primary"):
    salva_database_json()
    st.sidebar.success("Database JSON salvato con successo!")

# ── TASTO NUOVA GIORNATA ─────────────────────────────────────────────────────
st.sidebar.markdown("---")
st.sidebar.subheader("📅 Nuova Giornata")
st.sidebar.caption("Crea il tabellone del giorno successivo ereditando Stato e Furgone dall'ultima giornata archiviata. I Km Inizio vengono impostati pari ai Km Inizio dell'ultima giornata (il sistema aggiorna i Km Fine/Percorsi del giorno precedente in automatico).")

if st.sidebar.button("➕ NUOVA GIORNATA", use_container_width=True):
    storico_ref = st.session_state.storico_presenze.copy()
    df_anag     = st.session_state.anagrafica_corrieri.copy()

    if storico_ref.empty:
        # Nessuno storico: reset pulito dall'anagrafica
        df_nuovo = df_anag.copy()
        df_nuovo["STATO"]         = "Presente (Giro Fisso)"
        df_nuovo["GIRO_SUPPORTO"] = ""
        df_nuovo["MEZZO"]         = "Nessuno"
        df_nuovo["KM_INIZIO"]     = 0
        df_nuovo["NOTE"]          = ""
        st.session_state.stato_giornaliero = df_nuovo
        salva_database_json()
        st.sidebar.success("✅ Nuovo tabellone creato dall'anagrafica (nessuno storico disponibile).")
    else:
        # Trova la data più recente nello storico
        storico_ref["_data_dt"] = storico_ref["DATA"].apply(parse_data_ita)
        storico_valido = storico_ref[storico_ref["_data_dt"].notna()]

        if storico_valido.empty:
            st.sidebar.warning("⚠️ Nessuna data valida trovata nello storico.")
        else:
            data_max_dt   = storico_valido["_data_dt"].max()
            data_max_label = storico_valido[storico_valido["_data_dt"] == data_max_dt].iloc[0]["DATA"]
            righe_ref     = storico_valido[storico_valido["_data_dt"] == data_max_dt]

            # Costruisci il nuovo tabellone partendo dall'anagrafica
            df_nuovo = df_anag.copy()
            df_nuovo["STATO"]         = "Presente (Giro Fisso)"
            df_nuovo["GIRO_SUPPORTO"] = ""
            df_nuovo["MEZZO"]         = "Nessuno"
            df_nuovo["KM_INIZIO"]     = 0
            df_nuovo["NOTE"]          = ""

            # Recupera Km Inizio corrente dal tabellone attivo (= i km del giorno in corso)
            df_corrente = st.session_state.stato_giornaliero.copy()

            for idx in df_nuovo.index:
                cognome = df_nuovo.at[idx, "COGNOME"]
                nome    = df_nuovo.at[idx, "NOME"]

                # 1. Eredita STATO e MEZZO dall'ultima giornata archiviata
                match_st = righe_ref[
                    (righe_ref["COGNOME"] == cognome) &
                    (righe_ref["NOME"]    == nome)
                ]
                if not match_st.empty:
                    r = match_st.iloc[0]
                    df_nuovo.at[idx, "STATO"] = r.get("STATO", "Presente (Giro Fisso)")
                    df_nuovo.at[idx, "MEZZO"] = r.get("MEZZO", "Nessuno")
                    df_nuovo.at[idx, "GIRO_SUPPORTO"] = r.get("GIRO_SUPPORTO", "")

                # 2. KM_INIZIO nuova giornata = KM_INIZIO del tabellone corrente
                #    (cioè i km che erano stati inseriti stamattina, che diventano
                #     km fine della giornata che si sta chiudendo)
                match_corr = df_corrente[
                    (df_corrente["COGNOME"] == cognome) &
                    (df_corrente["NOME"]    == nome)
                ]
                if not match_corr.empty:
                    km_inizio_nuovo = int(match_corr.iloc[0].get("KM_INIZIO", 0) or 0)
                    df_nuovo.at[idx, "KM_INIZIO"] = km_inizio_nuovo

                    # 3. Aggiorna KM_FINE e KM_PERCORSI del giorno precedente nello storico
                    #    usando i KM_INIZIO della nuova giornata come KM_FINE di quella vecchia
                    mask = (
                        (st.session_state.storico_presenze["DATA"]    == data_max_label) &
                        (st.session_state.storico_presenze["COGNOME"] == cognome) &
                        (st.session_state.storico_presenze["NOME"]    == nome)
                    )
                    if mask.any():
                        km_i_prec = int(st.session_state.storico_presenze.loc[mask, "KM_INIZIO"].iloc[0] or 0)
                        st.session_state.storico_presenze.loc[mask, "KM_FINE"]     = km_inizio_nuovo
                        st.session_state.storico_presenze.loc[mask, "KM_PERCORSI"] = max(km_inizio_nuovo - km_i_prec, 0)

            st.session_state.stato_giornaliero = df_nuovo
            salva_database_json()
            st.sidebar.success(f"✅ Nuova giornata creata! Dati ereditati dal **{data_max_label}**. Km Fine di quella giornata aggiornati nello storico.")

# ── ESPORTA (visibili solo nella scheda Tabellone) ──────────
if scelta == "📋 Tabellone Presenze" and st.session_state.get("excel_sidebar_data") is not None:
    st.sidebar.markdown("---")
    st.sidebar.subheader("📤 Esporta Piano")
    st.sidebar.download_button(
        label="📥 Excel",
        data=st.session_state.excel_sidebar_data,
        file_name=st.session_state.excel_sidebar_nome,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
    st.sidebar.download_button(
        label="📥 PDF",
        data=st.session_state.pdf_sidebar_data,
        file_name=st.session_state.pdf_sidebar_nome,
        mime="application/pdf",
        use_container_width=True
    )

# ─────────────────────────────────────────────────────────────────────────────
# 1. TABELLONE PRESENZE
# ─────────────────────────────────────────────────────────────────────────────
if scelta == "📋 Tabellone Presenze":
    data_lavorazione = st.date_input("Data di lavorazione del Piano Presenze", datetime.today())
    giorno = data_lavorazione.day
    anno = data_lavorazione.year
    mese_testo = MESI_ITA[data_lavorazione.month]
    data_formato_personalizzato = f"{giorno} {mese_testo} {anno}"

    # FIX #4 – Ricalcola la lista furgoni ad ogni render così rispecchia sempre
    # lo stato attuale (anche dopo aggiunte/rimozioni nella scheda Furgoni)
    furgoni_attivi = st.session_state.furgoni[st.session_state.furgoni['DISPONIBILE'] != "GUASTO"]
    tutte_targhe = furgoni_attivi['TARGA'].tolist()

    # Calcola furgoni già assegnati nel tabellone corrente
    mezzi_in_uso = set(
        v for v in st.session_state.stato_giornaliero['MEZZO'].tolist()
        if v and v != "Nessuno"
    )
    targhe_libere    = [t for t in tutte_targhe if t not in mezzi_in_uso]
    targhe_assegnate = [t for t in tutte_targhe if t in mezzi_in_uso]

    # Menu a tendina con separatori visivi: prima i liberi, poi i già assegnati
    elenco_furgoni_tendina = (
        ["Nessuno"]
        + (["── NON ASSEGNATI ──"] if targhe_libere    else [])
        + targhe_libere
        + (["── GIÀ ASSEGNATI ──"] if targhe_assegnate else [])
        + targhe_assegnate
    )

    def salva_tabellone_giornaliero():
        if "editor_giornaliero_diretto" in st.session_state:
            edits = st.session_state["editor_giornaliero_diretto"]
            df_attuale = st.session_state.stato_giornaliero.copy()

            for row_idx, deltas in edits["edited_rows"].items():
                for col, val in deltas.items():
                    df_attuale.iat[row_idx, df_attuale.columns.get_loc(col)] = val

            # ── ANTI-DUPLICATO FURGONI ────────────────────────────────────────
            # Logica corretta:
            # 1. Identifica le righe appena modificate con un nuovo MEZZO
            # 2. Se quel mezzo era già presente su un'altra riga, libera QUELLA riga
            #    (così la nuova assegnazione viene sempre rispettata)
            righe_modificate_mezzo = {
                row_idx: deltas["MEZZO"]
                for row_idx, deltas in edits["edited_rows"].items()
                if "MEZZO" in deltas
            }

            for riga_nuova, nuovo_mezzo in righe_modificate_mezzo.items():
                if nuovo_mezzo in ("Nessuno", "── NON ASSEGNATI ──", "── GIÀ ASSEGNATI ──", ""):
                    continue
                # Cerca se lo stesso mezzo è presente su un'altra riga
                for idx, row in df_attuale.iterrows():
                    if idx != riga_nuova and row["MEZZO"] == nuovo_mezzo:
                        # Libera la riga che aveva il furgone in precedenza
                        df_attuale.iat[idx, df_attuale.columns.get_loc("MEZZO")] = "Nessuno"
                        nome_prev = f"{df_attuale.iat[idx, df_attuale.columns.get_loc('COGNOME')]} {df_attuale.iat[idx, df_attuale.columns.get_loc('NOME')]}"
                        st.info(f"ℹ️ Il furgone **{nuovo_mezzo}** è stato spostato da **{nome_prev}** al nuovo autista.")
                        break

            # Pulisci eventuali voci-separatore selezionate per errore
            col_mezzo = df_attuale.columns.get_loc("MEZZO")
            separatori = {"── NON ASSEGNATI ──", "── GIÀ ASSEGNATI ──"}
            for idx in range(len(df_attuale)):
                if df_attuale.iat[idx, col_mezzo] in separatori:
                    df_attuale.iat[idx, col_mezzo] = "Nessuno"

            st.session_state.stato_giornaliero = df_attuale

    n_righe = len(st.session_state.stato_giornaliero)
    altezza_griglia = max(300, n_righe * 35 + 38)  # 35px per riga + header 38px

    st.data_editor(
        st.session_state.stato_giornaliero,
        column_config={
            "COGNOME":       st.column_config.TextColumn("Cognome",  disabled=True),
            "NOME":          st.column_config.TextColumn("Nome",     disabled=True),
            "CELLULARE":     st.column_config.TextColumn("Cellulare", disabled=True),
            "GIRO_FISSO":    st.column_config.TextColumn("Giro Fisso", disabled=True),
            "STATO": st.column_config.SelectboxColumn(
                "Stato Presenza",
                options=["Presente (Giro Fisso)", "Supporto Altra Filiale", "Assente"],
                required=True,
                width="medium"
            ),
            "GIRO_SUPPORTO": st.column_config.TextColumn("Giro di Supporto / Filiale", width="medium"),
            "MEZZO": st.column_config.SelectboxColumn(
                "Furgone Assegnato (Targa)",
                options=elenco_furgoni_tendina,
                required=True,
                width="medium"
            ),
            "KM_INIZIO": st.column_config.NumberColumn("Km Inizio Giornata", min_value=0, step=1, width="small"),
            "NOTE": st.column_config.TextColumn("Note Operative", width="large")
        },
        hide_index=True,
        use_container_width=True,
        height=altezza_griglia,
        key="editor_giornaliero_diretto",
        on_change=salva_tabellone_giornaliero
    )

    # ── ARCHIVIA GIORNATA ─────────────────────────────────────────────────────
    st.markdown("---")
    col_arch1, col_arch2 = st.columns([2, 1])
    with col_arch1:
        st.markdown("#### 📁 Archivia giornata nello storico")
        st.caption("Salva la giornata corrente nello storico. I Km Percorsi verranno calcolati quando si crea la giornata successiva (Km Inizio successivo − Km Inizio attuale).")
    with col_arch2:
        if st.button("📁 ARCHIVIA GIORNATA", type="primary", use_container_width=True):
            df_oggi = st.session_state.stato_giornaliero.copy()
            data_str = data_formato_personalizzato

            # Costruisci righe dello storico
            # KM_FINE e KM_PERCORSI vengono calcolati quando si crea la giornata successiva
            nuove_righe = []
            for _, row in df_oggi.iterrows():
                km_i = int(row.get("KM_INIZIO", 0) or 0)
                nuove_righe.append({
                    "DATA":        data_str,
                    "COGNOME":     row["COGNOME"],
                    "NOME":        row["NOME"],
                    "GIRO_FISSO":  row["GIRO_FISSO"],
                    "STATO":       row["STATO"],
                    "MEZZO":       row.get("MEZZO", "Nessuno"),
                    "KM_INIZIO":   km_i,
                    "KM_FINE":     0,
                    "KM_PERCORSI": 0,
                    "NOTE":        row.get("NOTE", ""),
                })

            if nuove_righe:
                df_nuove = pd.DataFrame(nuove_righe)
                # Rimuovi eventuale archiviazione precedente della stessa data
                st_storico = st.session_state.storico_presenze
                if not st_storico.empty and "DATA" in st_storico.columns:
                    st_storico = st_storico[st_storico["DATA"] != data_str]
                st.session_state.storico_presenze = pd.concat(
                    [st_storico, df_nuove], ignore_index=True
                )
                salva_database_json()
                st.success(f"✅ Giornata del {data_str} archiviata! Usa **'➕ Nuova Giornata'** nella barra laterale per preparare il tabellone del giorno successivo.")
            else:
                st.warning("Nessun dato da archiviare.")

    df_correnti = st.session_state.stato_giornaliero
    # Garantisce colonne KM anche se df viene da una versione precedente del DB
    for _c, _d in [("KM_INIZIO", 0), ("MEZZO", "Nessuno"),
                   ("GIRO_SUPPORTO", ""), ("NOTE", ""), ("STATO", "Presente (Giro Fisso)"),
                   ("CELLULARE", ""), ("GIRO_FISSO", ""), ("COGNOME", ""), ("NOME", "")]:
        if _c not in df_correnti.columns:
            df_correnti[_c] = _d

    blocco1 = df_correnti[df_correnti['STATO'] == "Presente (Giro Fisso)"][
        ['COGNOME', 'NOME', 'CELLULARE', 'GIRO_FISSO', 'MEZZO', 'KM_INIZIO', 'NOTE']]
    blocco2 = df_correnti[df_correnti['STATO'] == "Supporto Altra Filiale"][
        ['COGNOME', 'NOME', 'CELLULARE', 'GIRO_SUPPORTO', 'MEZZO', 'KM_INIZIO', 'NOTE']]
    blocco3 = st.session_state.responsabili
    blocco4 = df_correnti[df_correnti['STATO'] == "Assente"][
        ['COGNOME', 'NOME', 'CELLULARE', 'NOTE']]

    # ── ESPORTAZIONE EXCEL ────────────────────────────────────────────────────
    def genera_excel_4_blocchi(data_label):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Crea il foglio con un DataFrame vuoto per avere il foglio disponibile
            pd.DataFrame().to_excel(writer, sheet_name='Piano Giornaliero', index=False)
            ws = writer.sheets['Piano Giornaliero']

            font_data_top  = Font(name='Calibri', size=11, bold=True)
            font_titolo    = Font(name='Calibri', size=11, bold=True)
            font_header    = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            fill_header    = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            fill_titolo    = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            border_grigio  = Border(
                left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),  bottom=Side(style='thin', color='D9D9D9')
            )
            allineamento_centro = Alignment(horizontal='center', vertical='center')

            ws.cell(row=1, column=2, value="PRESENZE CORRIERI").font        = font_data_top
            ws.cell(row=1, column=3, value=f"DATA: {data_label}").font      = font_data_top
            ws.cell(row=1, column=4, value="Filiale SDA PORTO D'ASCOLI").font = font_data_top

            def scrivi_blocco_excel(ws, df, titolo_blocco, riga_partenza):
                ws.cell(row=riga_partenza, column=1, value=titolo_blocco).font = font_titolo
                ws.cell(row=riga_partenza, column=1).fill = fill_titolo
                ws.row_dimensions[riga_partenza].height = 22
                for col_idx, col_name in enumerate(df.columns, start=1):
                    cell = ws.cell(row=riga_partenza + 1, column=col_idx, value=col_name)
                    cell.font      = font_header
                    cell.fill      = fill_header
                    cell.alignment = allineamento_centro
                    cell.border    = border_grigio
                ws.row_dimensions[riga_partenza + 1].height = 20
                curr_row = riga_partenza + 2
                for _, riga_dati in df.iterrows():
                    for col_idx, val in enumerate(riga_dati, start=1):
                        cell = ws.cell(row=curr_row, column=col_idx,
                                       value=str(val) if pd.notna(val) else "")
                        cell.border = border_grigio
                        if col_idx in [3, 4, 5]:
                            cell.alignment = allineamento_centro
                    ws.row_dimensions[curr_row].height = 18
                    curr_row += 1
                return curr_row + 2

            # FIX #7 – rimosso ws.delete_rows inutile; il foglio è appena creato
            prossima_riga = 3
            prossima_riga = scrivi_blocco_excel(ws, blocco1, "1° BLOCCO: CORRIERI CON NUMERO DI GIRO ASSOCIATO",     prossima_riga)
            prossima_riga = scrivi_blocco_excel(ws, blocco2, "2° BLOCCO: CORRIERI IN SUPPORTO ALTRA FILIALE",        prossima_riga)
            prossima_riga = scrivi_blocco_excel(ws, blocco3, "3° BLOCCO: NOMINATIVI RESPONSABILI PRESENTI",          prossima_riga)
            prossima_riga = scrivi_blocco_excel(ws, blocco4, "4° BLOCCO: CORRIERI ASSENTI",                          prossima_riga)

            for col in ws.columns:
                max_len    = max((len(str(cell.value or '')) for cell in col), default=0)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 13)

        return output.getvalue()

    # ── ESPORTAZIONE PDF ──────────────────────────────────────────────────────
    def genera_pdf_4_blocchi(data_label):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", "B", 14)
        pdf.cell(0, 10, "PIANO GIORNALIERO FLOTTA E PRESENZE CORRIERI", align="C")
        pdf.ln(6)
        pdf.set_font("Arial", "I", 11)
        pdf.cell(0, 10, f"Data di lavorazione: {data_label}", align="C")
        pdf.ln(12)

        def aggiungi_tabella_pdf(titolo, df):
            pdf.set_font("Arial", "B", 10)
            pdf.set_fill_color(225, 230, 240)
            pdf.cell(0, 7, titolo, fill=True)
            pdf.ln(9)
            pdf.set_font("Arial", "", 9)
            if df.empty:
                pdf.cell(0, 7, " Nessun record registrato in questo blocco.")
                pdf.ln(10)
                return
            pdf.set_fill_color(30, 75, 120)
            pdf.set_text_color(255, 255, 255)
            for col in df.columns:
                pdf.cell(31, 7, str(col), border=1, fill=True)
            pdf.ln(7)
            pdf.set_text_color(0, 0, 0)
            for _, riga in df.iterrows():
                for col in df.columns:
                    pdf.cell(31, 7, str(riga[col])[:16], border=1)
                pdf.ln(7)
            pdf.ln(4)

        aggiungi_tabella_pdf("1. CORRIERI CON GIRO ASSOCIATO",         blocco1)
        aggiungi_tabella_pdf("2. EVENTUALI CORRIERI IN SUPPORTO",       blocco2)
        aggiungi_tabella_pdf("3. NOMINATIVI RESPONSABILI PRESENTI",     blocco3)
        aggiungi_tabella_pdf("4. CORRIERI ASSENTI",                     blocco4)
        return bytes(pdf.output())

    st.session_state.excel_sidebar_data = genera_excel_4_blocchi(data_formato_personalizzato)
    st.session_state.pdf_sidebar_data   = genera_pdf_4_blocchi(data_formato_personalizzato)
    st.session_state.excel_sidebar_nome = f"Presenze {data_formato_personalizzato}.xlsx"
    st.session_state.pdf_sidebar_nome   = f"Presenze {data_formato_personalizzato}.pdf"



# ─────────────────────────────────────────────────────────────────────────────
# 2. SCHEDA ANAGRAFICA FURGONI
# ─────────────────────────────────────────────────────────────────────────────
elif scelta == "🚐 Anagrafica Furgoni":
    def salva_furgoni_edits():
        if "tabella_gestione_furgoni" in st.session_state:
            edits = st.session_state["tabella_gestione_furgoni"]
            df_attuale = st.session_state.furgoni.copy()
            for row_idx, deltas in edits["edited_rows"].items():
                for col, val in deltas.items():
                    df_attuale.iat[row_idx, df_attuale.columns.get_loc(col)] = val
            if edits["deleted_rows"]:
                df_attuale = df_attuale.drop(edits["deleted_rows"]).reset_index(drop=True)
            if edits["added_rows"]:
                for new_row in edits["added_rows"]:
                    riga_pulita = {col: new_row.get(col, "") for col in df_attuale.columns}
                    df_attuale = pd.concat([df_attuale, pd.DataFrame([riga_pulita])], ignore_index=True)
            st.session_state.furgoni = df_attuale

    st.data_editor(
        st.session_state.furgoni,
        num_rows="dynamic",
        column_config={
            "DISPONIBILE": st.column_config.SelectboxColumn(
                "Disponibilità", options=["SI", "NO", "GUASTO"], required=True
            )
        },
        use_container_width=True,
        key="tabella_gestione_furgoni",
        on_change=salva_furgoni_edits
    )

# ─────────────────────────────────────────────────────────────────────────────
# 3. SCHEDA ANAGRAFICA PERSONALE
# ─────────────────────────────────────────────────────────────────────────────
elif scelta == "👥 Anagrafica Personale":
    def aggiorna_anagrafica_corrieri():
        if "tabella_gestione_corrieri" in st.session_state:
            edits = st.session_state["tabella_gestione_corrieri"]
            df_attuale = st.session_state.anagrafica_corrieri.copy()

            for row_idx, deltas in edits["edited_rows"].items():
                for col, val in deltas.items():
                    df_attuale.iat[row_idx, df_attuale.columns.get_loc(col)] = val

            if edits["deleted_rows"]:
                df_attuale = df_attuale.drop(edits["deleted_rows"]).reset_index(drop=True)

            if edits["added_rows"]:
                for new_row in edits["added_rows"]:
                    riga_pulita = {col: new_row.get(col, "") for col in df_attuale.columns}
                    df_attuale = pd.concat([df_attuale, pd.DataFrame([riga_pulita])], ignore_index=True)

            st.session_state.anagrafica_corrieri = df_attuale

            # FIX #2 – NON rigenerare il tabellone giornaliero da zero:
            # aggiorna SOLO le colonne anagrafiche (COGNOME, NOME, CELLULARE, GIRO_FISSO)
            # preservando STATO, GIRO_SUPPORTO, MEZZO e NOTE già inseriti.
            df_giorno = st.session_state.stato_giornaliero.copy()

            # Allinea le righe esistenti
            for col in ["COGNOME", "NOME", "CELLULARE", "GIRO_FISSO"]:
                if col in df_attuale.columns and col in df_giorno.columns:
                    # Aggiorna solo le righe esistenti (per indice)
                    for idx in df_attuale.index:
                        if idx < len(df_giorno):
                            df_giorno.iat[idx, df_giorno.columns.get_loc(col)] = \
                                df_attuale.iat[idx, df_attuale.columns.get_loc(col)]

            # Aggiungi nuovi corrieri se l'anagrafica è cresciuta
            if len(df_attuale) > len(df_giorno):
                nuovi = df_attuale.iloc[len(df_giorno):].copy()
                nuovi["STATO"]         = "Presente (Giro Fisso)"
                nuovi["GIRO_SUPPORTO"] = ""
                nuovi["MEZZO"]         = "Nessuno"
                nuovi["KM_INIZIO"]     = 0
                nuovi["KM_FINE"]       = 0
                nuovi["NOTE"]          = ""
                df_giorno = pd.concat([df_giorno, nuovi], ignore_index=True)

            # Rimuovi righe se l'anagrafica è diminuita
            if len(df_attuale) < len(df_giorno):
                df_giorno = df_giorno.iloc[:len(df_attuale)].reset_index(drop=True)

            st.session_state.stato_giornaliero = df_giorno

    st.data_editor(
        st.session_state.anagrafica_corrieri,
        num_rows="dynamic",
        use_container_width=True,
        key="tabella_gestione_corrieri",
        on_change=aggiorna_anagrafica_corrieri
    )

    st.markdown("#### 👔 Responsabili Presenti")

    def aggiorna_responsabili():
        if "tabella_responsabili" in st.session_state:
            edits = st.session_state["tabella_responsabili"]
            df_r = st.session_state.responsabili.copy()
            for row_idx, deltas in edits["edited_rows"].items():
                for col, val in deltas.items():
                    df_r.iat[row_idx, df_r.columns.get_loc(col)] = val
            if edits["deleted_rows"]:
                df_r = df_r.drop(edits["deleted_rows"]).reset_index(drop=True)
            if edits["added_rows"]:
                for new_row in edits["added_rows"]:
                    riga_pulita = {col: new_row.get(col, "") for col in df_r.columns}
                    df_r = pd.concat([df_r, pd.DataFrame([riga_pulita])], ignore_index=True)
            st.session_state.responsabili = df_r

    st.data_editor(
        st.session_state.responsabili,
        num_rows="dynamic",
        use_container_width=True,
        key="tabella_responsabili",
        on_change=aggiorna_responsabili
    )


# ─────────────────────────────────────────────────────────────────────────────
# 4. STORICO PRESENZE & ANALISI FURGONI
# ─────────────────────────────────────────────────────────────────────────────
elif scelta == "📊 Storico & Furgoni":
    st.title("📊 Storico Presenze & Analisi Utilizzo Furgoni")

    # ── PANNELLO IMPORT EXCEL ─────────────────────────────────────────────────
    with st.expander("📥 Importa giornate da Excel", expanded=False):
        st.markdown(
            "Carica il file Excel compilato con le giornate mancanti. "
            "Le righe con la stessa data già presente nello storico verranno **sovrascritte**."
        )

        # Download template
        def _genera_template_bytes():
            import io as _io
            from openpyxl import Workbook as _WB
            from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side
            from openpyxl.worksheet.datavalidation import DataValidation as _DV
            from openpyxl.utils import get_column_letter as _gcl
            _corrieri = st.session_state.anagrafica_corrieri[["COGNOME","NOME","CELLULARE","GIRO_FISSO"]].values.tolist()
            _wb = _WB()
            _ws = _wb.active
            _ws.title = "DATI_DA_IMPORTARE"
            _cols = ["DATA","COGNOME","NOME","CELLULARE","GIRO_FISSO","STATO","MEZZO","KM_INIZIO","NOTE"]
            _larg = [22,18,14,15,12,26,14,12,30]
            _fb = _Fill("solid", start_color="1F4E78", end_color="1F4E78")
            _fg = _Fill("solid", start_color="F2F2F2", end_color="F2F2F2")
            _fy = _Fill("solid", start_color="FFF2CC", end_color="FFF2CC")
            _thin = _Side(style="thin", color="D9D9D9")
            _bord = _Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
            for c, (col, larg) in enumerate(zip(_cols, _larg), 1):
                cell = _ws.cell(row=1, column=c, value=col)
                cell.font = _Font(name="Calibri", bold=True, color="FFFFFF", size=11)
                cell.fill = _fb
                cell.alignment = _Align(horizontal="center")
                _ws.column_dimensions[_gcl(c)].width = larg
            _dv = _DV(type="list", formula1='"Presente (Giro Fisso),Supporto Altra Filiale,Assente"', showDropDown=False)
            _ws.add_data_validation(_dv)
            for r, (cogn, nome, cell_, giro) in enumerate(_corrieri, 2):
                _ws.row_dimensions[r].height = 17
                for c, val in enumerate(["", cogn, nome, cell_, giro, "Presente (Giro Fisso)", "Nessuno", 0, ""], 1):
                    cell = _ws.cell(row=r, column=c, value=val)
                    cell.font = _Font(name="Calibri", size=10)
                    cell.border = _bord
                    if c == 1: cell.fill = _fy
                    elif c in (2,3,4,5): cell.fill = _fg; cell.font = _Font(name="Calibri", size=10, color="555555")
                    elif c == 6: _dv.add(cell)
            _ws.freeze_panes = "A2"
            buf = _io.BytesIO()
            _wb.save(buf)
            return buf.getvalue()

        st.download_button(
            "📄 Scarica Template Excel",
            data=_genera_template_bytes(),
            file_name="template_import_presenze.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Scarica il template con tutti i corrieri già inseriti. Compila la colonna DATA e modifica STATO/MEZZO/KM dove necessario."
        )

        st.markdown("---")
        file_import = st.file_uploader(
            "Carica file Excel compilato (.xlsx)",
            type=["xlsx"],
            key="uploader_import_storico"
        )

        if file_import is not None:
            try:
                df_import = pd.read_excel(file_import, sheet_name="DATI_DA_IMPORTARE", dtype=str)
                df_import.columns = [c.strip().upper() for c in df_import.columns]

                # Rimuovi la riga di esempio (DATA vuota o uguale a "DATA")
                df_import = df_import[
                    df_import["DATA"].notna() &
                    (df_import["DATA"].str.strip() != "") &
                    (df_import["DATA"].str.strip().str.upper() != "DATA")
                ].copy()

                # Normalizza colonne numeriche
                for _nc in ["KM_INIZIO"]:
                    if _nc in df_import.columns:
                        df_import[_nc] = pd.to_numeric(df_import[_nc], errors="coerce").fillna(0).astype(int)

                # Assicura colonne mancanti
                for _mc, _mv in [("KM_FINE", 0), ("KM_PERCORSI", 0), ("NOTE", ""), ("MEZZO", "Nessuno"), ("STATO", "Presente (Giro Fisso)"), ("GIRO_FISSO", "")]:
                    if _mc not in df_import.columns:
                        df_import[_mc] = _mv

                df_import["DATA"] = df_import["DATA"].str.strip()

                date_trovate = sorted(df_import["DATA"].unique())
                n_righe = len(df_import)

                st.info(f"📊 **{n_righe} righe** trovate · **{len(date_trovate)} date**: {', '.join(date_trovate)}")

                col_imp1, col_imp2 = st.columns([1,2])
                with col_imp1:
                    if st.button("✅ IMPORTA NELLO STORICO", type="primary", use_container_width=True):
                        # Rimuovi dal storico attuale le date che stiamo importando
                        st_att = st.session_state.storico_presenze
                        if not st_att.empty and "DATA" in st_att.columns:
                            st_att = st_att[~st_att["DATA"].isin(date_trovate)]

                        # Allinea colonne
                        colonne_std = ["DATA","COGNOME","NOME","GIRO_FISSO","STATO","MEZZO","KM_INIZIO","KM_FINE","KM_PERCORSI","NOTE"]
                        for _c in colonne_std:
                            if _c not in df_import.columns:
                                df_import[_c] = 0 if _c in ("KM_INIZIO","KM_FINE","KM_PERCORSI") else ""

                        st.session_state.storico_presenze = pd.concat(
                            [st_att, df_import[colonne_std]], ignore_index=True
                        )
                        salva_database_json()
                        st.success(f"✅ Importate **{n_righe} righe** per le date: {', '.join(date_trovate)}. Database salvato.")
                        st.rerun()
                with col_imp2:
                    st.dataframe(df_import[["DATA","COGNOME","NOME","STATO","MEZZO","KM_INIZIO"]].head(10), hide_index=True, use_container_width=True)

            except Exception as _e:
                st.error(f"❌ Errore nella lettura del file: {_e}")

    storico = st.session_state.storico_presenze.copy()

    if storico.empty:
        st.info("📭 Nessuna giornata ancora archiviata. Vai al Tabellone Presenze e clicca **Archivia Giornata** al termine di ogni giorno.")
    else:
        # Assicura tipi numerici
        for _c in ["KM_INIZIO", "KM_FINE", "KM_PERCORSI"]:
            if _c in storico.columns:
                storico[_c] = pd.to_numeric(storico[_c], errors="coerce").fillna(0).astype(int)

        tab_storico, tab_furgoni = st.tabs(["📋 Storico Giornaliero", "🚐 Analisi per Furgone"])

        # ── TAB 1: STORICO GIORNALIERO ────────────────────────────────────────
        with tab_storico:
            st.subheader("Registro completo giornate archiviate")

            date_disponibili = sorted(storico["DATA"].unique(), reverse=True)
            col_f1, col_f2 = st.columns([2, 2])
            with col_f1:
                filtro_data = st.selectbox("Filtra per data", ["Tutte"] + date_disponibili)
            with col_f2:
                filtro_stato = st.selectbox("Filtra per stato", ["Tutti", "Presente (Giro Fisso)", "Supporto Altra Filiale", "Assente"])

            df_vis = storico.copy()
            if filtro_data != "Tutte":
                df_vis = df_vis[df_vis["DATA"] == filtro_data]
            if filtro_stato != "Tutti":
                df_vis = df_vis[df_vis["STATO"] == filtro_stato]

            st.dataframe(
                df_vis.reset_index(drop=True),
                use_container_width=True,
                hide_index=True,
                column_config={
                    "DATA":        st.column_config.TextColumn("Data"),
                    "COGNOME":     st.column_config.TextColumn("Cognome"),
                    "NOME":        st.column_config.TextColumn("Nome"),
                    "GIRO_FISSO":  st.column_config.TextColumn("Giro"),
                    "STATO":       st.column_config.TextColumn("Stato"),
                    "MEZZO":       st.column_config.TextColumn("Furgone"),
                    "KM_INIZIO":   st.column_config.NumberColumn("Km Inizio"),
                    "KM_FINE":     st.column_config.NumberColumn("Km Fine"),
                    "KM_PERCORSI": st.column_config.NumberColumn("Km Percorsi"),
                    "NOTE":        st.column_config.TextColumn("Note"),
                }
            )

            # Esporta storico Excel
            def esporta_storico_excel():
                out = io.BytesIO()
                with pd.ExcelWriter(out, engine="openpyxl") as writer:
                    df_vis.to_excel(writer, sheet_name="Storico", index=False)
                    ws = writer.sheets["Storico"]
                    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF")
                    hdr_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    for cell in ws[1]:
                        cell.font = hdr_font
                        cell.fill = hdr_fill
                    for col in ws.columns:
                        ml = max((len(str(c.value or "")) for c in col), default=0)
                        ws.column_dimensions[get_column_letter(col[0].column)].width = max(ml + 4, 12)
                return out.getvalue()

            st.download_button(
                "📥 Scarica Storico Excel",
                data=esporta_storico_excel(),
                file_name=f"Storico_Presenze.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # ── TAB 2: ANALISI PER FURGONE ────────────────────────────────────────
        with tab_furgoni:
            st.subheader("Utilizzo e km medi per furgone")

            # Solo righe con mezzo assegnato e km percorsi > 0
            df_fur = storico[
                (storico["MEZZO"] != "Nessuno") &
                (storico["MEZZO"].notna()) &
                (storico["MEZZO"] != "")
            ].copy()

            if df_fur.empty:
                st.info("Nessun dato di utilizzo furgoni trovato nello storico.")
            else:
                # ── Riepilogo tutti i furgoni ──
                st.markdown("##### 📊 Riepilogo generale")
                riepilogo = (
                    df_fur.groupby("MEZZO")
                    .agg(
                        Giorni_Utilizzo=("DATA", "nunique"),
                        Km_Totali=("KM_PERCORSI", "sum"),
                        Km_Medi_Giorno=("KM_PERCORSI", "mean"),
                        Autisti_Diversi=("COGNOME", "nunique"),
                    )
                    .reset_index()
                    .rename(columns={"MEZZO": "Furgone"})
                )
                riepilogo["Km_Medi_Giorno"] = riepilogo["Km_Medi_Giorno"].round(1)
                riepilogo = riepilogo.sort_values("Km_Totali", ascending=False)

                st.dataframe(
                    riepilogo,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Furgone":          st.column_config.TextColumn("Furgone (Targa)"),
                        "Giorni_Utilizzo":  st.column_config.NumberColumn("Giorni Utilizzo"),
                        "Km_Totali":        st.column_config.NumberColumn("Km Totali"),
                        "Km_Medi_Giorno":   st.column_config.NumberColumn("Km Medi/Giorno"),
                        "Autisti_Diversi":  st.column_config.NumberColumn("N° Autisti Diversi"),
                    }
                )

                # ── Expander per ogni furgone ──────────────────────────────────
                st.markdown("##### 🔍 Dettaglio giornaliero per furgone")

                col_ord1, col_ord2 = st.columns([2, 1])
                with col_ord1:
                    ordinamento_sel = st.selectbox(
                        "Ordina furgoni per",
                        options=["Targa (A→Z)", "Giorni Utilizzo (↓)", "Km Totali (↓)"],
                        key="ordinamento_expander"
                    )
                with col_ord2:
                    direzione_asc = st.checkbox("Inverti ordine", value=False, key="direzione_expander")

                _map_ord = {
                    "Targa (A→Z)":          ("Furgone",         True),
                    "Giorni Utilizzo (↓)":  ("Giorni_Utilizzo", False),
                    "Km Totali (↓)":        ("Km_Totali",       False),
                }
                _col_ord, _asc_default = _map_ord[ordinamento_sel]
                _asc_finale = (not _asc_default) if direzione_asc else _asc_default
                riepilogo_ordinato = riepilogo.sort_values(_col_ord, ascending=_asc_finale)

                st.caption(
                    f"{'↑' if _asc_finale else '↓'} Ordinato per **{ordinamento_sel.split(' (')[0]}** "
                    f"— {len(riepilogo_ordinato)} furgoni con utilizzo registrato"
                )

                for _, row_riepilogo in riepilogo_ordinato.iterrows():
                    targa       = row_riepilogo["Furgone"]
                    giorni_uso  = int(row_riepilogo["Giorni_Utilizzo"])
                    km_tot      = int(row_riepilogo["Km_Totali"])
                    km_medi     = float(row_riepilogo["Km_Medi_Giorno"])
                    n_autisti   = int(row_riepilogo["Autisti_Diversi"])

                    label_exp = (
                        f"🚐 **{targa}** — "
                        f"{giorni_uso} gg utilizzo · "
                        f"{km_tot:,} km totali · "
                        f"{km_medi} km/gg · "
                        f"{n_autisti} autist{'a' if n_autisti == 1 else 'i'}"
                    )
                    with st.expander(label_exp, expanded=False):
                        df_det = df_fur[df_fur["MEZZO"] == targa].copy()
                        # Ordina per data
                        df_det["_data_dt"] = df_det["DATA"].apply(parse_data_ita)
                        df_det = df_det.sort_values("_data_dt", ascending=False).drop(columns=["_data_dt"], errors="ignore")

                        cols_da_mostrare = [c for c in ["DATA","COGNOME","NOME","STATO","KM_INIZIO","KM_FINE","KM_PERCORSI","NOTE"] if c in df_det.columns]
                        st.dataframe(
                            df_det[cols_da_mostrare].reset_index(drop=True),
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "DATA":        st.column_config.TextColumn("Data"),
                                "COGNOME":     st.column_config.TextColumn("Cognome"),
                                "NOME":        st.column_config.TextColumn("Nome"),
                                "STATO":       st.column_config.TextColumn("Stato"),
                                "KM_INIZIO":   st.column_config.NumberColumn("Km Inizio"),
                                "KM_FINE":     st.column_config.NumberColumn("Km Fine"),
                                "KM_PERCORSI": st.column_config.NumberColumn("Km Percorsi"),
                                "NOTE":        st.column_config.TextColumn("Note"),
                            }
                        )

                        # Scarica dettaglio furgone
                        def _esporta_det(df_export, targa_exp):
                            out = io.BytesIO()
                            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                                df_export.to_excel(writer, sheet_name=targa_exp[:31], index=False)
                                ws = writer.sheets[targa_exp[:31]]
                                hdr_font = Font(name="Calibri", bold=True, color="FFFFFF")
                                hdr_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                                for cell in ws[1]:
                                    cell.font = hdr_font
                                    cell.fill = hdr_fill
                                for col in ws.columns:
                                    ml = max((len(str(c.value or "")) for c in col), default=0)
                                    ws.column_dimensions[get_column_letter(col[0].column)].width = max(ml + 4, 12)
                            return out.getvalue()

                        st.download_button(
                            f"📥 Scarica dettaglio {targa}",
                            data=_esporta_det(df_det[cols_da_mostrare], targa),
                            file_name=f"Furgone_{targa}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_furgone_{targa}"
                        )
